import datetime
import time
import threading
import requests
import os
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
import schedule



url = "https://api.openligadb.de/getmatchdata/bl1/2025"

def excel():
    if os.path.exists("Spiele.xlsx"):
        wb = load_workbook("Spiele.xlsx")
    else:
        wb = Workbook()
        wb.save("Spiele.xlsx")
    return wb


def Spiele(url):
    response = requests.get(url)
    data = response.json()
    for i in data:
      spieltag = i ["group"]["groupName"]
      home = i ["team1"] ["teamName"]
      away = i ["team2"] ["teamName"]
      match_date_time = i ["matchDateTime"]
      id = i ["matchID"]
      if spieltag in wb.sheetnames:
         ws = wb[spieltag]
         ws.append([home, away, match_date_time, id])
      else:
         ws = wb.create_sheet(title=spieltag, index = 0)
         ws.append(["Heim", "Gast", "Zeitpunt", "ID"])
         ws.append([home, away, match_date_time, id])
    wb.save("Spiele.xlsx")
     


wb = excel()
Spiele(url)