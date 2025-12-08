import tweepy
from dotenv import load_dotenv
import os
import datetime
from openpyxl import load_workbook
import time

def keys():
    time.sleep(2)
    load_dotenv()
    bearer_key = os.getenv("BEARER_KEY")
    api_key = os.getenv("API_KEY")
    api_key_secret = os.getenv("API_KEY_SECRET")
    access_token = os.getenv("ACCESS_TOKEN")
    access_token_secret = os.getenv("ACCESS_TOKEN_SECRET")
    client = tweepy.Client(bearer_key,api_key,api_key_secret,access_token,access_token_secret)
    auth = tweepy.OAuth1UserHandler(api_key, api_key_secret, access_token, access_token_secret)
    api = tweepy.API(auth)
    return client, api#, bearer_key, api_key, api_key_secret, access_token, access_token_secret 
    


def tweet(post):
    print("Working dir:", os.getcwd())
    client, _ = keys()
    print(post)
    response = client.create_tweet(text=post)
    print("POst")
    print(response)


def tweet_table():
    path = "Screenshots_table/table.png"
    client, api = keys()
    media_id = api.media_upload(path).media_id
    with open("Spieltag.txt", "r") as f:
        matchday = f.read().strip()
    matchday, _ = matchday.split(".")
    text = f"Bundesliga table after Matchday {matchday}! 📊\n\nWhere does your team rank? Discuss ⬇️\nLike and follow for more!\n\n#Bundesliga #BuLi #table #football #goals #sport #standings"
    client.create_tweet(text=text, media_ids=[media_id])


def mdprinter(matchday):
    md,irr = matchday.split(".")
    wb = load_workbook("Spiele.xlsx")
    ws = wb[matchday]
    

    text = f"⚽Bundesliga MD{md}\n"
    current_day = None
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        home = row[0]
        away = row[1]
        date = row[2]

        date_obj = datetime.datetime.strptime(date, "%Y-%m-%dT%H:%M:%S")
        weekday = date_obj.strftime("%a")  # Kürzer: Mon, Tue, ...
        time_str = date_obj.strftime("%H:%M")
        
        if weekday != current_day:
            text += f"\n{weekday}\n"
            current_day = weekday

        text += f"{time_str} {home} 🆚 {away}\n"
        
        # Check if we are over 280 characters
        if len(text) > 280:
            text = text[:277] + "..."
            break

    return text


def mprinter(data, bl):
    text = "🕐FT\n"

    result = data["matchResults"][1]

    home = data["team1"] ["shortName"]
    home_col,home_short = bl[data["team1"]["teamName"]].split("| ")
    away = data["team2"] ["shortName"]
    away_col,away_short = bl[data["team2"]["teamName"]].split("| ")
    home_goals = result["pointsTeam1"]
    away_goals = result["pointsTeam2"]
    number_goals = int(home_goals) + int(away_goals)


   
    text += f"{home_col}{home} {home_goals} - {away_goals} {away}{away_col} \n\n"



    for e in data["goals"]:
        min = e ["matchMinute"]
        scorer = e ["goalGetterName"]
        text += f"\n⚽ {min}' {scorer}"
        if e["isPenalty"]:
            text += " (P)"
        if e["isOwnGoal"]:
            text += " (OG)"


    text += f"\n\n#{home_short}{away_short} #bundesliga #football #goals"
    return text
    


    
