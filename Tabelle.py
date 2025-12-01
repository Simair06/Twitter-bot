import openpyxl
from openpyxl.drawing.image import Image
import os
import requests
import excel2img

# Beispiel-Dictionary, jetzt mit lokalem Pfad zum Logo
logos = {
    "FC Bayern München" : "logos/Bayern.png",
    "RB Leipzig" : "logos/Leipzig.png",
    "Borussia Dortmund" : "logos/BVB.png",
    "VfB Stuttgart" : "logos/Stuttgart.png" ,
    "Bayer 04 Leverkusen" : "logos/Leverkusen.png",
    "TSG Hoffenheim" : "logos/Hoffenheim.png",
    "Eintracht Frankfurt" : "logos/Frankfurt.png",
    "SV Werder Bremen" : "logos/Bremen.png",
    "1. FC Köln" : "logos/Köln.png",
    "SC Freiburg" : "logos/Freiburg.png",
    "1. FC Union Berlin" : "logos/Union.png",
    "Borussia Mönchengladbach" : "logos/Gladbach.png",
    "Hamburger SV" : "logos/HSV.png",
    "VfL Wolfsburg" : "logos/Wolfsburg.png",
    "FC Augsburg" : "logos/Augsburg.png",
    "FC St. Pauli" : "logos/Pauli.png",
    "1. FSV Mainz 05" : "logos/Mainz.png",
    "1. FC Heidenheim 1846" : "logos/Heidenheim.png"
}

def table_api():
    url = "https://api.openligadb.de/getbltable/bl1/2025"
    data = requests.get(url)
    data = data.json()
    table = []
    for team in data:
        info = []
        wdl = f"{team["won"]}-{team["draw"]}-{team["lost"]}"
        info = [team["teamName"], team["matches"], wdl, team["goalDiff"], team["points"]]
        table.append(info)
    return table 






def excel_writer(table):
    # Name der vorhandenen Excel-Datei
    excel_file = "table.xlsx"

    # Arbeitsmappe öffnen
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    for img in list(ws._images):  
        ws._images.remove(img)


    
    # Daten einfügen
    for i, team in enumerate(table, start = 1): 
        cells = {"1": 5,"2": 7,"3": 9,"4": 11,"5": 13,"6": 15,"7": 17,"8": 19,"9": 21,"10": 23,"11": 25,"12": 27,"13": 29,"14": 31,"15": 33,"16": 35,"17": 37,"18": 39}


        cell = cells[str(i)]

        # Spalten H–L füllen
        ws[f"H{cell}"] = team[0]
        ws[f"I{cell}"] = team[1]
        ws[f"J{cell}"] = team[2]
        ws[f"K{cell}"] = team[3]
        ws[f"L{cell}"] = team[4]

        # Logo aus lokalem Ordner einfügen
        logo_path = logos[team[0]]
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width = 20
            img.height = 20
            cell_ref = f"G{cell}"
            img.anchor = cell_ref
            
          
            ws.add_image(img)
        else:
            print(f"Logo nicht gefunden: {logo_path}")
        




    # Spaltenbreite anpassen
    ws.column_dimensions["G"].width = 9
    ws.column_dimensions["H"].width = 30

    # Speichern
    wb.save(excel_file)
    print(f"Excel-Datei '{excel_file}' erfolgreich aktualisiert ab Zeile 4!")


    output_file = "Screenshots_table/table.png"
    sheet_name = "Tabelle1"
    excel_range = "E2:M41"
    excel2img.export_img(excel_file, output_file, sheet_name, excel_range)







def main():
    table = table_api()
    excel_writer(table)
    

main()





