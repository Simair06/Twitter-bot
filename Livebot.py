import datetime
import requests
from Tweeter import tweet, mdprinter, mprinter, tweet_table
from openpyxl import load_workbook
import Tabelle as table

#https://www.openligadb.de/

current_matches_url = "https://api.openligadb.de/getmatchdata/bl1"
matchday_url = "https://api.openligadb.de/getcurrentgroup/bl1"
 

print(datetime.datetime.now())

bundesliga_vereine = {
    "FC Bayern München": "🔴⚪ | FCB",
    "Borussia Dortmund": "⚫🟡 | BVB",
    "RB Leipzig": "🔴⚪ | RBL",
    "Bayer 04 Leverkusen": "🔴⚫ | B04",
    "Eintracht Frankfurt": "⚫🔴 | SGE",
    "Borussia Mönchengladbach": "🟢⚫ | BMG",
    "VfB Stuttgart": "🔴⚪ | VfB",
    "1. FC Köln": "🔴⚪ | FC",
    "1. FC Union Berlin": "🔴⚪ | FCU",
    "1. FSV Mainz 05": "🔴⚪ | M05",
    "FC Augsburg": "🔴🟢 | FCA",
    "SC Freiburg": "⚫⚪ | SCF",
    "TSG Hoffenheim": "🔵⚪ | TSG",
    "SV Werder Bremen": "🟢⚪ | SVW",
    "VfL Wolfsburg": "🟢⚪ | WOB",
    "Hamburger SV": "🔵⚪ | HSV",
    "FC St. Pauli": "🟤⚪ | FCSP",
    "1. FC Heidenheim 1846": "🔴🔵 | FCH"
}

def api(url):
    response = requests.get(url)
    data = response.json()
    return data


def get_matchday():
    matchday = api(matchday_url)
    matchday = matchday["groupName"]
    with open("Spieltag.txt", "r") as f:
        current_content = f.read().strip()
    if current_content != matchday:
        with open("Spieltag.txt", "w") as f:
            f.write(matchday)
            post = mdprinter(matchday)
            tweet(post)
    return matchday
              

def live():
    with open("Spieltag.txt", "r") as f:
        matchday = f.read().strip()
    today = []
    wb = load_workbook("Spiele.xlsx")
    ws = wb[matchday]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        date, time = row[2].split("T")
        match_date = datetime.datetime.strptime(date, "%Y-%m-%d").date()
        if match_date == datetime.date.today():
            today.append({"id" : row[3], "time" : time })
    #today.append({"id" :  , "time" : "10:00:00" }) # Spiele die noch nicht gepostet wurden 
    print(today) 
    return today
       

def pruefe_job(bl):
    jetzt = datetime.datetime.now()
    jetzt = jetzt
    spiele = live()
  
    
    if len(spiele) > 0:
        for spiel in spiele:
            id = spiel["id"]
            data = requests.get(f"https://api.openligadb.de/getmatchdata/{id}")
            data = data.json()
            with open("posted.txt", "r") as f:
                content = f.read()
            if (data["matchIsFinished"]) and (str(id) not in content):
                    text = mprinter(data,bl)
                    tweet(text)
                    with open("posted.txt", "a") as f:
                        f.write(f"{str(id)}\n")


def txtdel():
    with open("posted.txt", "w") as f:
        f.write("")


def tablewriter():
    table.main()
    tweet_table()
    
    





def main():
    get_matchday()
    pruefe_job(bundesliga_vereine)
    now = datetime.datetime.now()
    if now.weekday() == 0 and now.hour >= 8 and now.hour < 14:  #0 ist Montag
        tablewriter()


    
    

main()





